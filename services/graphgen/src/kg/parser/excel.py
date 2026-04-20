"""Excel parser (.xlsx) using openpyxl."""

from __future__ import annotations

from pathlib import Path

from kg.parser import BaseParser, ParsedChunk, ParsedDocument


class ExcelParser(BaseParser):
    extensions = ("xlsx", "xls")
    mime_types = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",)

    def parse(self, path: Path) -> ParsedDocument:
        try:
            import openpyxl  # type: ignore[import]
        except ImportError as exc:
            raise ImportError("openpyxl is required: pip install openpyxl") from exc

        doc = ParsedDocument.from_path(path, self.mime_types[0])
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        chunks: list[ParsedChunk] = []
        position = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None and str(c).strip()]
                if cells:
                    rows.append(" | ".join(cells))

            # Group rows into 50-row chunks
            chunk_rows = 50
            for i in range(0, max(1, len(rows)), chunk_rows):
                chunk_text = "\n".join(rows[i : i + chunk_rows])
                if chunk_text.strip():
                    chunks.append(ParsedChunk.make(doc.doc_id, position, chunk_text, [sheet_name]))
                    position += 1

        doc.chunks = chunks
        return doc
